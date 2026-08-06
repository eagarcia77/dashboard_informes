#!/usr/bin/env python3
from __future__ import annotations
import argparse, html, re
from collections import Counter
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

FOLDER=Path(__file__).resolve().parent
OUTPUT=FOLDER/'ucan_course_development_dashboard.html'

def clean(v): return '' if v is None else ' '.join(str(v).replace('\n',' ').split())
def money(v): return f'${v:,.0f}'

def header(ws):
    for r in range(1,min(ws.max_row,30)+1):
        vals=[clean(ws.cell(r,c).value).lower() for c in range(1,ws.max_column+1)]
        if 'recurso' in vals and 'curso' in vals: return r,{v:i+1 for i,v in enumerate(vals) if v}
    return None

def column(headers,*needles):
    for h,i in headers.items():
        if all(n in h for n in needles): return i
    return None

def status_group(s):
    s=s.lower()
    if any(k in s for k in ('pagada','pagado','certific','complet')): return 'Pagado / certificado'
    if any(k in s for k in ('aprob','acept','compromiso','contrato')): return 'Aprobado / comprometido'
    if any(k in s for k in ('propuesta','nuevo','new','pendiente')): return 'Propuesta / pendiente'
    return 'Sin estatus' if not s else 'En proceso / seguimiento'

def parse(path):
    wb=load_workbook(path,data_only=True); records=[]; years=Counter(); finances=Counter(); gaps=[]
    for ws in wb.worksheets:
        if ws.title.lower().startswith('faltan'):
            for row in ws.iter_rows(min_row=2,values_only=True):
                vals=list(row)+[None,None,None]
                if clean(vals[0]): gaps.append((clean(vals[0]),clean(vals[2])))
            continue
        found=header(ws)
        if not found: continue
        hr,h=found; rc=column(h,'recurso'); cc=column(h,'curso'); sc=column(h,'estatus'); tc=column(h,'total','comprometido')
        for r in range(hr+1,ws.max_row+1):
            resource=clean(ws.cell(r,rc).value) if rc else ''; course=clean(ws.cell(r,cc).value) if cc else ''
            if not course or course.lower() in {'curso','total'}: continue
            status=clean(ws.cell(r,sc).value) if sc else ''; raw=ws.cell(r,tc).value if tc else 0
            try: total=float(raw or 0)
            except (TypeError,ValueError): total=0
            m=re.match(r'([A-Za-z]{3,6})',course)
            rec={'year':ws.title,'resource':resource or 'No identificado','course':course,'program':m.group(1).upper() if m else 'OTROS','group':status_group(status),'total':total}
            records.append(rec); years[ws.title]+=1; finances[ws.title]+=total
    return records,years,finances,gaps

def bars(items,formatter=lambda x:str(int(x))):
    if not items: return '<p>No hay datos disponibles.</p>'
    mx=max(v for _,v in items) or 1; h=len(items)*48+34; out=[f'<svg class="chart" viewBox="0 0 900 {h}">']
    for i,(label,val) in enumerate(items):
        y=24+i*48; fill=max(3,val/mx*525); label=html.escape(label[:45]+('…' if len(label)>45 else ''))
        out += [f'<text x="10" y="{y+19}" class="svg-label">{label}</text>',f'<rect x="335" y="{y}" width="525" height="26" rx="13" class="bar-bg"/>',f'<rect x="335" y="{y}" width="{fill:.1f}" height="26" rx="13" class="bar-fill"/>',f'<text x="{min(341+fill,840):.1f}" y="{y+19}" class="svg-value">{html.escape(formatter(val))}</text>']
    return '\n'.join(out+['</svg>'])

def table(rows,headers,numeric=()):
    head=''.join(f'<th>{html.escape(x)}</th>' for x in headers); body=[]
    for row in rows:
        cells=''.join(f'<td{" class=\"num\"" if i in numeric else ""}>{html.escape(str(v))}</td>' for i,v in enumerate(row))
        body.append(f'<tr>{cells}</tr>')
    return f'<table><thead><tr>{head}</tr></thead><tbody>{"".join(body)}</tbody></table>'

def render(records,years,finances,gaps,source):
    resources={r['resource'] for r in records if r['resource']!='No identificado'}; programs=Counter(r['program'] for r in records); statuses=Counter(r['group'] for r in records)
    total=sum(r['total'] for r in records); paid=sum(r['total'] for r in records if r['group']=='Pagado / certificado'); unique={r['course'] for r in records}
    year_items=sorted(years.items()); program_items=programs.most_common(12); status_items=statuses.most_common(); finance_items=sorted(finances.items())
    css=':root{--green:#007B5F;--yellow:#FED141;--ink:#1f2937;--muted:#64748b;--bg:#f4f7f6;--card:#fff;--line:#dbe4e1;--soft:#e7f3ef;--shadow:0 18px 45px rgba(15,23,42,.10)}*{box-sizing:border-box}body{margin:0;font-family:Arial,sans-serif;background:radial-gradient(circle at top left,rgba(254,209,65,.30),transparent 35%),radial-gradient(circle at top right,rgba(0,123,95,.20),transparent 35%),var(--bg);color:var(--ink)}header{background:linear-gradient(135deg,var(--green),#003c31);color:#fff;padding:34px 28px 32px}.header-inner,main,footer{max-width:1180px;margin:auto}.eyebrow{display:inline-block;color:var(--yellow);font-weight:700;text-transform:uppercase;font-size:12px}h1{font-size:clamp(30px,5vw,56px);margin:18px 0 8px}.subtitle{color:#e5f4ef;font-size:18px}.meta{display:flex;flex-wrap:wrap;gap:10px;margin-top:20px}.pill{background:rgba(255,255,255,.12);padding:9px 12px;border-radius:999px;font-size:13px}main{margin-top:-24px;padding:0 18px 60px}.grid{display:grid;grid-template-columns:repeat(6,1fr);gap:16px}.kpi,.section{background:#fff;border:1px solid var(--line);border-radius:20px;box-shadow:var(--shadow)}.kpi{padding:20px;min-height:150px}.kpi-value{font-size:34px;font-weight:800;color:var(--green)}.kpi-title{font-weight:800;margin-top:8px}.kpi-note,.lead{color:var(--muted);font-size:13px}.section{margin-top:22px;padding:24px}.section h2{margin:0 0 6px;color:#0f3f35}.two-col{display:grid;grid-template-columns:1.1fr .9fr;gap:22px}table{width:100%;border-collapse:collapse}th{background:var(--green);color:#fff;text-align:left;padding:12px}td{padding:12px;border-top:1px solid var(--line);font-size:14px}.num{text-align:right;font-weight:800}.chart{width:100%;background:#fff;border:1px solid var(--line);border-radius:18px;padding:12px}.svg-label,.svg-value{font-size:13px;fill:#334155;font-weight:700}.bar-bg{fill:#e8f0ed}.bar-fill{fill:#007B5F}.note-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}.note{background:var(--soft);padding:16px;border-radius:18px}.note strong{display:block;margin-bottom:6px}footer{padding:0 18px 36px;color:var(--muted);font-size:12px;text-align:center}@media(max-width:980px){.grid{grid-template-columns:repeat(2,1fr)}.two-col{grid-template-columns:1fr}}@media(max-width:560px){.grid{grid-template-columns:1fr}}'
    kpis=[(len(unique),'Cursos identificados','Cursos únicos registrados'),(len(records),'Registros de desarrollo','Filas válidas consolidadas'),(len(resources),'Recursos docentes','Personas identificadas'),(len(programs),'Áreas de cursos','Prefijos académicos'),(money(total),'Cartera financiera','Total comprometido'),(len(gaps),'Programas con brechas','Hoja Faltan Year 3')]
    cards=''.join(f'<div class="kpi"><div class="kpi-value">{v}</div><div class="kpi-title">{t}</div><div class="kpi-note">{n}</div></div>' for v,t,n in kpis)
    yr=[[y,c,money(finances[y])] for y,c in year_items]; pr=[[p,c] for p,c in program_items]; st=[[s,c,f'{c/len(records)*100:.1f}%'] for s,c in status_items]
    return f'''<!doctype html><html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>UCAN Dashboard de Desarrollo de Cursos</title><style>{css}</style></head><body><header><div class="header-inner"><span class="eyebrow">UCAN / PPOHA · Recinto de San Germán</span><h1>Dashboard de Desarrollo de Cursos</h1><p class="subtitle">Panel ejecutivo generado automáticamente desde el Excel de seguimiento.</p><div class="meta"><span class="pill">Preparado por: Dr. Eduardo Augusto García Rodríguez</span><span class="pill">Fuente: {html.escape(source.name)}</span><span class="pill">Actualizado: {datetime.now().strftime('%B %d, %Y')}</span></div></div></header><main><section class="grid">{cards}</section><section class="section"><h2>Resumen por año</h2><div class="two-col"><div>{bars([(k,float(v)) for k,v in year_items])}</div><div>{table(yr,['Año','Registros','Total'],(1,2))}</div></div></section><section class="section"><h2>Distribución por área</h2><div class="two-col"><div>{bars([(k,float(v)) for k,v in program_items])}</div><div>{table(pr,['Área / prefijo','Cursos'],(1,))}</div></div></section><section class="section"><h2>Estado operacional</h2><div class="two-col"><div>{bars([(k,float(v)) for k,v in status_items])}</div><div>{table(st,['Clasificación','Registros','Porcentaje'],(1,2))}</div></div></section><section class="section"><h2>Resumen financiero</h2><div class="two-col"><div>{bars([(k,float(v)) for k,v in finance_items],money)}</div><div class="note-grid"><div class="note"><strong>Total registrado</strong>{money(total)}</div><div class="note"><strong>Pagado / certificado</strong>{money(paid)}</div><div class="note"><strong>Balance operacional</strong>{money(total-paid)}</div></div></div></section><section class="section"><h2>Brechas de desarrollo</h2>{table(gaps,['Programa académico','Cursos faltantes o que requieren decisión'])}</section></main><footer>Dashboard generado automáticamente por GitHub Actions.</footer></body></html>'''

def choose(explicit):
    if explicit:
        p=Path(explicit)
        return p if p.is_absolute() else Path.cwd()/p
    preferred=FOLDER/'reporte_desarrollo_cursos.xlsx'
    if preferred.exists(): return preferred
    files=sorted(p for p in FOLDER.glob('*.xlsx') if not p.name.startswith('~$'))
    if not files: raise FileNotFoundError('No hay archivos .xlsx en Informe_Desarrollo_Cursos.')
    return files[-1]

def main():
    parser=argparse.ArgumentParser(); parser.add_argument('--input'); args=parser.parse_args(); source=choose(args.input)
    records,years,finances,gaps=parse(source)
    if not records: raise RuntimeError('El Excel no contiene filas de cursos reconocibles.')
    OUTPUT.write_text(render(records,years,finances,gaps,source),encoding='utf-8')
    print(f'Dashboard actualizado: {OUTPUT}; registros: {len(records)}')
if __name__=='__main__': main()
